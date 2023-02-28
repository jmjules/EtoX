# Benötigte Libraries zum Auslesen der Excel Files und dem Erstellen des GUIs
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import *

#
def convert_to_xml():
    # Öffnen des Excel Files und Angabe des Pfads im UI
    file_path = "data.xlsx" #filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    wb = load_workbook(file_path, rich_text=True)
    ws = wb.active

    # Erstellen der XML Struktur
    root = ET.Element('transliterateUnits')

    #start bei 2, um die Sprachen zu überspringen
    
    for row in ws.iter_rows(2):
        row_el = ET.SubElement(root, 'unit')
        
        for cell in row:
            tag = ws.cell(1, cell.column).value #nimmt Tag-benennung aus erster Zeile jeder Spalte
            
            #Operationen je nach Zellwert: None -> leeres element | Zahl -> ELement mit Zahl als String | textBlock -> Element mit Text
            if cell.value == None:
                cell_el = ET.SubElement(row_el, str(tag))
            elif isinstance(cell.value, int):
                cell_el = ET.SubElement(row_el, str(tag))
                cell_el.text = str(cell.value)
            else:
                #Rich-Texte werden in objekten gespeichert und müssen separat behandelt werden
                if isinstance(cell.value[0], object):
                    tempText = ""
                    for part in cell.value:
                        if hasattr(part, "font"):
                            align = part.font.vertAlign
                            if align != None:
                                insert = part.text.replace("&","&amp;" ).replace("<", "&lt;").replace(">", "&gt;")
                                tempText += f"<{align}>{insert}</{align}>"
                            else:
                                tempText += part.text.replace("&","&amp;" ).replace("<", "&lt;").replace(">", "&gt;")
                        else:
                            tempText += part.replace("&","&amp;" ).replace("<", "&lt;").replace(">", "&gt;")
                    lastTempText = f'<{tag}>{tempText}</{tag}>'
                    cell_el = ET.fromstring(lastTempText)
                    row_el.append(cell_el)
                else:
                    cell_el = ET.SubElement(row_el, str(tag))
                    cell_el.text = str(cell.value)
    #Edge case: wenn alles in der Zelle auf die selbe weise formatiert ist, funktioniert es nicht. ->es wird immer eine "normale formatierung zusammen mit einer "besonderen erwartet (Das war aber ja auch nicht bei den vorgaben dabei)       
    #print("---------------------")

    #Das XML leserlich machen
    ET.indent(root)
    # Schreiben des XML Files sowie angeben des Speicherorts
    tree = ET.ElementTree(root)
    #file_path = filedialog.asksaveasfilename(defaultextension=".xml", filetypes=[("XML files", "*.xml")])
    #tree.writeStandalone(file_path, encoding='utf-8', xml_declaration=True)
    tree.writeStandalone("data.xml", encoding='utf-8', xml_declaration=True) #just for testing ->delete
    #messagebox.showinfo("Information", "Datei wurde erfolgreich konvertiert.")

root = tk.Tk()
root.title("Excel to XML Konverter")
root.geometry("300x150")

# Buttons
convert_button = tk.Button(root, text="Excel to XML konvertieren", command=convert_to_xml, height=2, bg='#4F7942', fg='#FFFFFF')
convert_button.place(relx=0.5, rely=0.5, anchor="center")

#root.mainloop()

convert_to_xml()


